import os
import re
import pandas as pd
from os.path import join as path
import numpy as np

# path1 = '/home/user/luoshengjie/code'
# sample = 'target.log'

# # match = re.match(r"$target*", os.)
# # print(match)
# path2 = path(path1,'{}', sample).format("*")
# # os.system('ls {fastq_path}/{sample}*.fastq.gz'.format(fastq_path=fastq_path,sample=sample))
# print(path2)
# with open(path2,'r') as f:
#     for line in f.readline():
#         print(line)
# os.system(f"""ls -lh {path2}""")

a = np.array([2,4,5,2])

print(a)

# !/usr/bin/python/
# -*- coding:utf-8 -*-
# author:hanqiulian
# date: 2022/06/21

import re
import click
import openpyxl
import pandas as pd
from os.path import join as path
from openpyxl.styles import Font, PatternFill


def conbine_arg_info(arg_path, sa, tb_arg, drugname_dict):
    mtb_result_file = path(arg_path, sa + '_argsite.txt')
    df = pd.read_csv(mtb_result_file, sep='\t')
    # 去除没有耐药信息的结果
    mtb_result = df[df['InterestingRegion'] != '-']
    # 删除部分列，以使结果更简洁
    # mtb_result=mtb_result.drop(labels=['Insindex',"CovFor","CovRev", "Qual20", "Freq"],axis=1)
    header = list(mtb_result.columns)
    header_dict = dict([element, i] for i, element in enumerate(header))
    # print(header_dict)
    mtb_result = mtb_result.values.tolist()

    # 去除同义突变结果
    for i in mtb_result[:]:
        amino_mutation_info = re.findall("[a-zA-Z]+", i[header_dict['Subst']].split()[0])
        if len(amino_mutation_info) == 2 and amino_mutation_info[0] == amino_mutation_info[1]:
            mtb_result.remove(i)

    # mtb_result.insert(mtb_result.shape[1],'drug_info','-')
    # 追加耐药翻译信息
    header.append('drug_info')
    [i.append(drugname_dict[i[header_dict['InterestingRegion']]]) if i[header_dict[
        'InterestingRegion']] in drugname_dict else i.append('-') for i in mtb_result]

    # 追加correct列，判别rpoB密码子落在509-81=428   534-81=453与否，如果是，则加81
    # header.append('correction')
    # [i.append(i[header_dict['Subst']][0:3] + str(int(re.search("\d+",i[header_dict['Subst']]).group())+81) +  i[header_dict['Subst']][6:9])   if i[-1]=='利福平' and re.search("\d+",i[header_dict['Subst']]) and 453>int(re.search("\d+",i[header_dict['Subst']]).group())>428   else i.append('-') for i in mtb_result ]

    # 追加correct列，判别rpoB密码子位置加81
    # header.append('correction')
    # [i.append(i[header_dict['Subst']][0:3] + str(int(re.search("\d+",i[header_dict['Subst']]).group())+81) +  i[header_dict['Subst']][6:9])   if i[-1]=='利福平' and re.search("\d+",i[header_dict['Subst']])  else i.append('-') for i in mtb_result ]

    # 方法2.不追加correct列，判别rpoB密码子位置加81，直接替换原来的
    for i in mtb_result:
        if i[-1] == '利福平' and re.search("\d+", i[header_dict['Subst']]):
            i[header_dict['Subst']] = re.sub(str(int(re.search("\d+", i[header_dict['Subst']]).group())),
                                             str(int(re.search("\d+", i[header_dict['Subst']]).group()) + 81),
                                             i[header_dict['Subst']])

    # 追加核酸突变结果
    header.append('核酸突变结果')
    for i in mtb_result:
        if re.search("\d+", i[header_dict['Subst']]) and re.search("/", i[
            header_dict['Subst']]):  # 判断是否有氨基酸位置及碱基突变前后信息 Subst信息格式为：His285Arg (cac/cGc)
            amino_pos = re.search("\d+", i[header_dict['Subst']]).group()
            mutation_info = i[header_dict['Subst']].split('/')[1]
            gene_pos = int(amino_pos) * 3 - 2 + mutation_info.index(re.search('[ATCG]', mutation_info).group())
            i.append(str(gene_pos) + i[header_dict['Ref']] + '>' + i[header_dict['Allel']])
        else:
            i.append('-')

    # 追加配置表信息
    header.extend(['物种', '药物', '耐药水平'])
    config_header = list(tb_arg.columns)
    config_header_dict = dict([element, i] for i, element in enumerate(config_header))
    tb_arg = tb_arg.values.tolist()
    for i in mtb_result:
        flag = 0
        for j in tb_arg:
            #			if i[-2] != '-' and i[header_dict['Subst']] != '-' and i[header_dict['GeneName']] == j[config_header_dict['基因名']]    and i[-2]==j[config_header_dict['密码子突变']]:   # 有校正值则识别校正值是否和配置表一致
            #				i.extend([j[config_header_dict['物种']], j[config_header_dict['药物']], j[config_header_dict['耐药水平']]])
            #				flag=1
            #			elif i[-2] == '-' and i[header_dict['Subst']] != '-' and i[header_dict['GeneName']] == j[config_header_dict['基因名']]    and i[header_dict['Subst']].split()[0]==j[config_header_dict['密码子突变']]:   # 没有校正值则识别Subst里的氨基酸突变信息和配置表是否一致
            #				i.extend([j[config_header_dict['物种']], j[config_header_dict['药物']], j[config_header_dict['耐药水平']]])
            #				flag=1

            if i[header_dict['Subst']] != '-' and i[header_dict['GeneName']] == j[config_header_dict['基因名']] and \
                    i[header_dict['Subst']].split()[0] == j[config_header_dict['密码子突变']] and len(
                    i) == 19:  # 校正值已经取消，rpoB的Subst内容直接被校正
                i.extend([j[config_header_dict['物种']], j[config_header_dict['药物']], j[config_header_dict['耐药水平']]])
                flag = 1

        if flag == 0:
            i.extend(['-', '-', '低'])

    # 追加样本编号
    header.insert(0, '样本编号')
    [i.insert(0, sa) for i in mtb_result]
    return header, mtb_result


def merge(total_result, samples):
    # 二位数组转dataframe
    columns = total_result.pop(0)
    df = pd.DataFrame(total_result, columns=columns)
    # 只取特定列
    df1 = df[['#Pos', 'GeneName', 'drug_info']]
    # 对dataframe按特定列排序，行号重置，去重
    df1 = df1.sort_values(by='#Pos', ascending=True).reset_index(drop=True).drop_duplicates()
    # 增加多列，并设置初始值
    for sa in samples:
        df1.insert(df1.shape[1], sa, '-')

    for index, row in df1.iterrows():
        # 提取在该位置有突变的样本列表
        sa_list = df[df['#Pos'] == row['#Pos']]['样本编号'].tolist()
        for i in sa_list:
            # df1.iloc[index,df1.columns.tolist().index(i)] = 'y'   #给dataframe的某行某列赋值，索引为数值
            df1.loc[index, i] = 'Y'  # 给dataframe的某行某列赋值，索引为数值
    return df1

def work(file):
    file = '/home/test/b.txt'
    alist = file.split('/')
    for i in alist:
        print(i**2,'this is a fault')
    return

# df1.to_csv('./test.txt',sep='\t')

def mark(ws, samples, colors, row_num):
    for row in ws.iter_rows(min_row=2, max_row=row_num):
        row[0].fill = PatternFill("solid", fgColor=colors[samples.index(row[0].value)])
        if row[1].value == 1673425:
            row[1].font = Font(color="FFFF0000")


if __name__ == '__main__':
    @click.command()
    @click.option('-s', '--samples_list',
                  help='MG20C01986-4,MG20C01986-5,MG20C02545-4,MG20C02545-5,MG20C02549-4,MG20C02549-5',
                  default='MT22C00759,MT22C00760')
    @click.option('-b', '--batch', help='TPNB500481_0014', default='TPNB500481_0429')
    @click.option('-a', '--arg_config', help='/home/pipeline/target_pathogen1/config/arg_sites_info.xlsx',
                  default='/home/pipeline/target_pathogen1/config/arg_sites_info.xlsx')
    @click.option('-p1', '--project_path', help='/home/clinical_projects/tNGS', default='/home/clinical_projects/tNGS')
    @click.option('-o', '--out_file', help='arg_result.xlsx', default='arg_result.xlsx')
    def m(samples_list, batch, arg_config, project_path, out_file):
        tb_arg = pd.read_excel(arg_config, engine="openpyxl")
        # print(tb_arg)
        '''
                  物种     药物   基因名 Gene position stop      密码子突变 密码子突变结果 耐药水平
                0     结核分枝杆菌    利福平  rpoB                436  Val146Phe       -    低
                1     结核分枝杆菌    利福平  rpoB                508  Val170Phe       -    低
                2     结核分枝杆菌    利福平  rpoB                  -  Val176Phe       -    低
                3     结核分枝杆菌    利福平  rpoB                  -  Pro206Arg       -    低
                4     结核分枝杆菌    利福平  rpoB                  -  Tyr314Cys       -    低
                ..       ...    ...   ...                ...        ...     ...  ...
                595  非结核分枝杆菌  氨基糖苷类   rrs               1409          -       -    高
                596  非结核分枝杆菌  氨基糖苷类   rrs               1409          -       -    高
                597  非结核分枝杆菌  氨基糖苷类   rrs               1491          -       -    高
                598  非结核分枝杆菌  氨基糖苷类   rrs               1491          -       -    高
                599  非结核分枝杆菌  氨基糖苷类   rrs               1491          -       -    高
        '''
        drugname_dict = {'fluoroquinolones (FQ)': '氟喹诺酮类', 'streptomycin (SM)': '链霉素', 'ethambutol (EMB)': '乙胺丁醇',
                         'ethionamid (ETH)': '乙硫异烟胺', 'rifampicin (RMP)': '利福平', 'linezolid (LZD)': '利奈唑胺',
                         'isoniazid (INH)': '异烟肼', 'pyrazinamide (PZA)': '吡嗪酰胺'}
        arg_path = path(project_path, batch, 'argsite_300')
        arg_result_file = path(project_path, batch, out_file)
        contant = ''
        flag = 0
        total_result = []
        wb = openpyxl.Workbook()
        for sa, snum in zip(samples_list.split(','), range(len(samples_list.split(',')))):
            header, mtb_result = conbine_arg_info(arg_path, sa, tb_arg, drugname_dict)
            ws = wb.create_sheet(sa, snum)
            ws.append(header)
            [ws.append(i) for i in mtb_result]
            if flag == 0:
                total_result.extend([header])
                total_result.extend(mtb_result)

                flag = 1
            else:
                total_result.extend(mtb_result)
        # print(total_result)
        ws = wb.create_sheet(title='merge1', index=0)
        [ws.append(i) for i in total_result]
        colors = ["FFFFDEAD", "FFE0FFFF", "FFD8BFD8"] * 40  # navajowhite，lightcyan，thistle 即纳瓦霍白，浅青色，蓟

        # 获取有耐药结果的数据清单
        sas = []
        for i in total_result[1:]:
            if i[0] not in sas:
                sas.append(i[0])
        mark(ws, sas, colors, len(total_result))

        df1 = merge(total_result, samples_list.split(','))
        ws = wb.create_sheet(title='merge0', index=0)
        ws.append(df1.columns.tolist())
        [ws.append(i) for i in df1.values.tolist()]

        out_file = path(project_path, batch, out_file)
        # 结果输出到文本文件
        # with open(out_file,'w') as f:
        # total_result = [[str(ii) for ii in i] for i in total_result ]
        # f.write(''.join('\t'.join(i)+'\n' for i in total_result))

        # 结果输出到xlsx
        wb.save(out_file)


    m()
